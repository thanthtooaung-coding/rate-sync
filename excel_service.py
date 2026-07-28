"""Excel persistence for multi-source exchange-rate records."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import EXCEL_PATH, ensure_directories
from models import RateRecord, SourceKind
from parser import ExchangeRate

AVG_SHEET_NAME: str = "Avg Rate"

AVG_HEADERS: tuple[str, ...] = (
    "Date",
    "Time",
    "Avg Buy 5Lakh",
    "Avg Buy 10Lakh",
    "Avg Sell",
    "Avg Sell Alt",
    "Sources",
)

SOURCE_HEADERS: tuple[str, ...] = (
    "Date",
    "Time",
    "Buy 5Lakh",
    "Buy 10Lakh",
    "Sell",
    "Sell Alt",
    "Message ID",
    "Raw Message",
)

MESSAGE_ID_COLUMN: int = 7
_INVALID_SHEET_CHARS: re.Pattern[str] = re.compile(r"[\\/*?:\[\]]")


def sheet_name_for_source(source: SourceKind, source_key: str) -> str:
    """Build an Excel sheet title for a Telegram channel or Facebook page.

    Args:
        source: Platform kind.
        source_key: Channel username or Facebook page id/name.

    Returns:
        A safe sheet title (max 31 characters).
    """
    safe: str = _INVALID_SHEET_CHARS.sub("_", source_key.strip().lstrip("@"))
    if source is SourceKind.FACEBOOK:
        candidates: tuple[str, ...] = (
            f"FB {safe}'s Exchange Rate",
            f"FB {safe}'s Rates",
            f"FB {safe}",
            safe,
        )
    else:
        candidates = (
            f"{safe}'s Exchange Rate",
            f"{safe}'s Rates",
            safe,
        )

    for title in candidates:
        if len(title) <= 31:
            return title
    prefix: str = "FB " if source is SourceKind.FACEBOOK else ""
    return f"{prefix}{safe}"[:31]


def storage_key(source: SourceKind, source_key: str) -> str:
    """Build a unique in-memory key for a source feed."""
    return f"{source.value}:{source_key.strip().lstrip('@').lower()}"


@dataclass(frozen=True)
class AverageRate:
    """Averaged rates across feeds that currently have data."""

    buy_5_lakh: float
    buy_10_lakh: float
    sell: float
    sell_alt: float | None
    sources_used: int


@dataclass
class ExcelService:
    """Persist rates into Avg Rate + Telegram sheets + Facebook sheets."""

    telegram_channels: list[str]
    facebook_pages: list[str] = field(default_factory=list)
    path: Path = EXCEL_PATH
    _known_ids: dict[str, set[str]] = field(default_factory=dict, init=False)
    _latest_rates: dict[str, ExchangeRate] = field(default_factory=dict, init=False)

    def __post_init__(self) -> None:
        """Ensure workbook/sheets exist and load known message IDs."""
        if not self.telegram_channels and not self.facebook_pages:
            raise ValueError("At least one Telegram channel or Facebook page is required")
        ensure_directories()
        self._ensure_workbook()
        self._known_ids = self._load_existing_message_ids()
        self._latest_rates = self._load_latest_rates()

    @property
    def _feed_specs(self) -> list[tuple[SourceKind, str]]:
        """Return configured feeds in sheet order: Telegram then Facebook."""
        specs: list[tuple[SourceKind, str]] = [
            (SourceKind.TELEGRAM, channel) for channel in self.telegram_channels
        ]
        specs.extend(
            (SourceKind.FACEBOOK, page) for page in self.facebook_pages
        )
        return specs

    def _create_workbook(self) -> None:
        """Create Avg Rate first, then Telegram sheets, then Facebook sheets."""
        workbook: Workbook = Workbook()
        default_sheet: Worksheet = workbook.active
        workbook.remove(default_sheet)

        avg_sheet: Worksheet = workbook.create_sheet(title=AVG_SHEET_NAME, index=0)
        avg_sheet.append(list(AVG_HEADERS))

        for source, key in self._feed_specs:
            sheet: Worksheet = workbook.create_sheet(
                title=sheet_name_for_source(source, key)
            )
            sheet.append(list(SOURCE_HEADERS))

        workbook.save(self.path)

    def _row_headers(self, sheet: Worksheet) -> tuple[object | None, ...]:
        """Return the header tuple for a sheet."""
        return tuple(
            cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        )

    def _workbook_schema_ok(self, workbook: Workbook) -> bool:
        """Validate Avg Rate + configured source sheet schemas."""
        if AVG_SHEET_NAME not in workbook.sheetnames:
            return False
        avg_headers: tuple[object | None, ...] = self._row_headers(
            workbook[AVG_SHEET_NAME]
        )
        if avg_headers[: len(AVG_HEADERS)] != AVG_HEADERS:
            return False

        for source, key in self._feed_specs:
            title: str = sheet_name_for_source(source, key)
            if title not in workbook.sheetnames:
                continue
            headers: tuple[object | None, ...] = self._row_headers(workbook[title])
            if headers[: len(SOURCE_HEADERS)] != SOURCE_HEADERS:
                return False
        return True

    def _ensure_avg_first(self, workbook: Workbook) -> None:
        """Move Avg Rate sheet to index 0."""
        if AVG_SHEET_NAME in workbook.sheetnames:
            workbook.move_sheet(
                AVG_SHEET_NAME,
                offset=-workbook.sheetnames.index(AVG_SHEET_NAME),
            )

    def _ensure_sheet_order(self, workbook: Workbook) -> None:
        """Reorder sheets: Avg Rate, Telegram feeds, Facebook feeds."""
        self._ensure_avg_first(workbook)
        desired: list[str] = [AVG_SHEET_NAME] + [
            sheet_name_for_source(source, key) for source, key in self._feed_specs
        ]
        for index, title in enumerate(desired):
            if title not in workbook.sheetnames:
                continue
            current_index: int = workbook.sheetnames.index(title)
            if current_index != index:
                workbook.move_sheet(title, offset=index - current_index)

    def _ensure_workbook(self) -> None:
        """Create the workbook and ensure all expected sheets exist."""
        if not self.path.exists():
            self._create_workbook()
            return

        workbook = load_workbook(self.path)
        if not self._workbook_schema_ok(workbook):
            workbook.close()
            backup: Path = self.path.with_suffix(".bak.xlsx")
            if backup.exists():
                backup.unlink()
            self.path.replace(backup)
            self._create_workbook()
            return

        if AVG_SHEET_NAME not in workbook.sheetnames:
            avg_sheet: Worksheet = workbook.create_sheet(title=AVG_SHEET_NAME, index=0)
            avg_sheet.append(list(AVG_HEADERS))

        for source, key in self._feed_specs:
            title: str = sheet_name_for_source(source, key)
            if title not in workbook.sheetnames:
                sheet: Worksheet = workbook.create_sheet(title=title)
                sheet.append(list(SOURCE_HEADERS))

        self._ensure_sheet_order(workbook)
        workbook.save(self.path)
        workbook.close()

    def _title_to_storage_key(self) -> dict[str, str]:
        """Map sheet titles to storage keys for configured feeds."""
        return {
            sheet_name_for_source(source, key): storage_key(source, key)
            for source, key in self._feed_specs
        }

    def _load_existing_message_ids(self) -> dict[str, set[str]]:
        """Load message IDs already stored per source sheet."""
        workbook = load_workbook(self.path, read_only=True, data_only=True)
        known: dict[str, set[str]] = {
            storage_key(source, key): set() for source, key in self._feed_specs
        }
        title_map: dict[str, str] = self._title_to_storage_key()

        for sheet in workbook.worksheets:
            feed_key: str | None = title_map.get(sheet.title)
            if feed_key is None:
                continue
            for row in sheet.iter_rows(
                min_row=2,
                min_col=MESSAGE_ID_COLUMN,
                max_col=MESSAGE_ID_COLUMN,
                values_only=True,
            ):
                value = row[0]
                if value is None:
                    continue
                known[feed_key].add(str(value))

        workbook.close()
        return known

    def _load_latest_rates(self) -> dict[str, ExchangeRate]:
        """Load the newest rate row from each configured source sheet."""
        workbook = load_workbook(self.path, read_only=True, data_only=True)
        latest: dict[str, ExchangeRate] = {}

        for source, key in self._feed_specs:
            title: str = sheet_name_for_source(source, key)
            if title not in workbook.sheetnames:
                continue
            sheet: Worksheet = workbook[title]
            last_rate: ExchangeRate | None = None
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    buy_5 = int(row[2])  # type: ignore[arg-type]
                    buy_10 = int(row[3])  # type: ignore[arg-type]
                    sell = int(row[4])  # type: ignore[arg-type]
                except (TypeError, ValueError, IndexError):
                    continue
                sell_alt: int | None
                try:
                    sell_alt = int(row[5]) if row[5] not in (None, "") else None
                except (TypeError, ValueError):
                    sell_alt = None
                last_rate = ExchangeRate(
                    buy_5_lakh=buy_5,
                    buy_10_lakh=buy_10,
                    sell=sell,
                    sell_alt=sell_alt,
                )
            if last_rate is not None:
                latest[storage_key(source, key)] = last_rate

        workbook.close()
        return latest

    def has_message(
        self,
        source: SourceKind,
        source_key: str,
        message_id: str,
    ) -> bool:
        """Check whether a message ID is already stored for a feed."""
        key: str = storage_key(source, source_key)
        return message_id in self._known_ids.get(key, set())

    def compute_average(self) -> AverageRate | None:
        """Average the latest rates across all feeds that have data."""
        rates: list[ExchangeRate] = list(self._latest_rates.values())
        if not rates:
            return None

        count: int = len(rates)
        sell_alts: list[int] = [
            rate.sell_alt for rate in rates if rate.sell_alt is not None
        ]
        return AverageRate(
            buy_5_lakh=round(sum(rate.buy_5_lakh for rate in rates) / count, 2),
            buy_10_lakh=round(sum(rate.buy_10_lakh for rate in rates) / count, 2),
            sell=round(sum(rate.sell for rate in rates) / count, 2),
            sell_alt=(
                round(sum(sell_alts) / len(sell_alts), 2) if sell_alts else None
            ),
            sources_used=count,
        )

    def write_average_snapshot(self, as_of: datetime | None = None) -> bool:
        """Append current cross-source average to the Avg Rate sheet."""
        average: AverageRate | None = self.compute_average()
        if average is None:
            return False

        stamp: datetime = as_of or datetime.now()
        workbook = load_workbook(self.path)
        if AVG_SHEET_NAME not in workbook.sheetnames:
            avg_sheet: Worksheet = workbook.create_sheet(title=AVG_SHEET_NAME, index=0)
            avg_sheet.append(list(AVG_HEADERS))
        else:
            avg_sheet = workbook[AVG_SHEET_NAME]
            self._ensure_avg_first(workbook)

        avg_sheet.append(
            [
                stamp.strftime("%Y-%m-%d"),
                stamp.strftime("%H:%M:%S"),
                average.buy_5_lakh,
                average.buy_10_lakh,
                average.sell,
                average.sell_alt if average.sell_alt is not None else "",
                average.sources_used,
            ]
        )
        workbook.save(self.path)
        workbook.close()
        return True

    def append_record(self, record: RateRecord, *, update_average: bool = False) -> bool:
        """Append a rate record to the matching source sheet if new."""
        key: str = storage_key(record.source, record.source_key)
        if self.has_message(record.source, record.source_key, record.message_id):
            return False

        workbook = load_workbook(self.path)
        title: str = sheet_name_for_source(record.source, record.source_key)
        if title not in workbook.sheetnames:
            sheet: Worksheet = workbook.create_sheet(title=title)
            sheet.append(list(SOURCE_HEADERS))
            self._ensure_sheet_order(workbook)
        else:
            sheet = workbook[title]

        sheet.append(
            [
                record.message_date.strftime("%Y-%m-%d"),
                record.message_date.strftime("%H:%M:%S"),
                record.rates.buy_5_lakh,
                record.rates.buy_10_lakh,
                record.rates.sell,
                record.rates.sell_alt if record.rates.sell_alt is not None else "",
                record.message_id,
                record.raw_message,
            ]
        )
        workbook.save(self.path)
        workbook.close()

        self._known_ids.setdefault(key, set()).add(record.message_id)
        self._latest_rates[key] = record.rates

        if update_average:
            self.write_average_snapshot(record.message_date)
        return True

    def append_many(self, records: Iterable[RateRecord]) -> tuple[int, int]:
        """Append multiple records, skipping duplicates."""
        inserted: int = 0
        skipped: int = 0
        for record in records:
            if self.append_record(record):
                inserted += 1
            else:
                skipped += 1
        return inserted, skipped
